#!/usr/bin/env python3
"""Regenerate galaxy/data.json from the YC market-map workbook(s).

Run from the yc_map1 folder after updating the spreadsheet:
    python3 galaxy/build_data.py

Base company records (sector / layer / orientation / customer / status / ...)
come from yc_market_map.xlsx. The B2A flag ("Serves agents") is overlaid from
the corrected master sheet if it is present, matched by (company, batch). The
overlay is purely additive: it sets `sa` on the 71 agent-serving companies and
changes nothing else, so classifications stay exactly as the base sheet has them.
"""
import json
import os

# openpyxl chokes on the float tabRatio these workbooks carry (Google Sheets
# export artifact) — coerce floats to int where an int is expected.
import openpyxl.descriptors.base as _b
_orig = _b._convert
def _patched(expected_type, value):
    try:
        return _orig(expected_type, value)
    except TypeError:
        if expected_type is int:
            return int(float(value))
        raise
_b._convert = _patched

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, '..', 'yc_market_map.xlsx')
OUT = os.path.join(HERE, 'data.json')

# The corrected master sheet carries the "Serves agents (B2A)" column. It lives
# outside the repo (it holds funding data); look in a few likely spots.
B2A_CANDIDATES = [
    os.path.join(HERE, '..', 'YC_market_map_corrected.xlsx'),
    os.path.expanduser('~/Downloads/YC_market_map_corrected.xlsx'),
]

SEASON = {'Winter': 0, 'Spring': 1, 'Summer': 2, 'Fall': 3}

def batch_key(b):
    season, year = b.split()
    return (int(year), SEASON[season])

wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb['Companies']
rows = ws.iter_rows(values_only=True)
hdr = [str(h).strip() if h else '' for h in next(rows)]
idx = {h: i for i, h in enumerate(hdr) if h}

def get(r, name):
    i = idx.get(name)
    if i is None or i >= len(r) or r[i] is None:
        return ''
    s = str(r[i]).strip()
    return '' if not s or s.startswith('=') else s

raw = []
for r in rows:
    name = get(r, 'Company')
    batch = get(r, 'Batch')
    if not name or not batch:
        continue
    raw.append((name, batch, r))

batches = sorted({b for _, b, _ in raw}, key=batch_key)
bidx = {b: i for i, b in enumerate(batches)}


def load_b2a():
    """Return a set of (company, batch) flagged 'Serves agents (B2A)' = Yes."""
    for path in B2A_CANDIDATES:
        if not os.path.exists(path):
            continue
        w = openpyxl.load_workbook(path, read_only=True)
        s = w['Companies']
        it = s.iter_rows(values_only=True)
        h = [str(x).strip() if x else '' for x in next(it)]
        ci = {name: i for i, name in enumerate(h) if name}
        col = ci.get('Serves agents (B2A)')
        nm, bt = ci.get('Company'), ci.get('Batch')
        if col is None or nm is None or bt is None:
            continue
        out = set()
        for r in it:
            if r[nm] is None:
                continue
            if str(r[col]).strip().lower() == 'yes':
                out.add((str(r[nm]).strip(), str(r[bt]).strip()))
        print(f'B2A overlay: {len(out)} companies from {os.path.basename(path)}')
        return out
    print('B2A overlay: corrected sheet not found — skipping (no `sa` flags)')
    return set()


b2a = load_b2a()

companies = []
for name, batch, r in raw:
    rec = {
        'n': name, 'b': bidx[batch],
        'o': get(r, 'One-liner'), 'w': get(r, 'Website'),
        'h': get(r, 'HQ location'), 'st': get(r, 'YC status') or 'Active',
        'ai': get(r, 'AI?') or 'Non-AI', 'ly': get(r, 'Layer'),
        'hv': get(r, 'Horizontal/Vertical'), 'se': get(r, 'Sector'),
        'su': get(r, 'Sub-sector'), 'fn': get(r, 'Function (if horizontal)'),
        'cu': get(r, 'Customer'), 'mo': get(r, 'Business model'),
    }
    if (name, batch) in b2a:
        rec['sa'] = 1
    companies.append(rec)

with open(OUT, 'w') as f:
    json.dump({'batches': batches, 'companies': companies}, f, separators=(',', ':'))

n_b2a = sum(1 for c in companies if c.get('sa'))
print(f'Wrote {len(companies)} companies across {len(batches)} batches '
      f'({n_b2a} flagged B2A) to {OUT}')
